import sqlite3
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from src.etl.loader import load_peer_groups

# -----------------------------
# Load Peer Groups
# -----------------------------

peer_df = load_peer_groups()

print("Peer Groups:", peer_df.shape)

# -----------------------------
# Load Financial Ratios
# -----------------------------

conn = sqlite3.connect("db/nifty100.db")

ratio_df = pd.read_sql(
    "SELECT * FROM financial_ratios",
    conn
)

conn.close()

print("Financial Ratios:", ratio_df.shape)

# -----------------------------
# Load Peer Percentiles
# -----------------------------

conn = sqlite3.connect("db/nifty100.db")

percentile_df = pd.read_sql(
    "SELECT * FROM peer_percentiles",
    conn
)

conn.close()

print("Peer Percentiles:", percentile_df.shape)

# -----------------------------
# Merge Peer Groups
# -----------------------------

report_df = ratio_df.merge(

    peer_df[
        [
            "company_id",
            "peer_group_name",
            "is_benchmark"
        ]
    ],

    on="company_id",

    how="left"

)

report_df["peer_group_name"] = (

    report_df["peer_group_name"]

    .fillna("No peer group assigned")

)

print(report_df.head())

# -----------------------------
# Create Workbook
# -----------------------------

wb = Workbook()

wb.remove(wb.active)

# -----------------------------
# Create One Sheet Per Peer Group
# -----------------------------

peer_groups = sorted(

    report_df[
        report_df["peer_group_name"] != "No peer group assigned"
    ]["peer_group_name"].unique()

)

print("\nPeer Groups Found:")

for group in peer_groups:
    print(group)

print("\nTotal Peer Groups:", len(peer_groups))

for group in peer_groups:

    wb.create_sheet(
        title=group[:31]
    )

print("\nWorkbook Sheets:")

print(wb.sheetnames)

# -----------------------------
# Columns for Report
# -----------------------------

report_columns = [

    "company_id",
    "year",

    "return_on_equity_pct",
    "roce_percentage",
    "net_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",

    "free_cash_flow_cr",
    "cash_from_operations_cr",

    "earnings_per_share",
    "book_value_per_share",
    "dividend_payout_ratio_pct",

    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",

    "composite_quality_score",
    "sector_relative_score",

    # Percentile Columns

    "ROE",
    "ROCE",
    "Net Profit Margin",
    "Debt to Equity",
    "Interest Coverage",
    "Asset Turnover",
    "Revenue CAGR 5Y",
    "PAT CAGR 5Y",
    "EPS CAGR 5Y"

]

# -----------------------------
# Pivot Percentile Table
# -----------------------------

percentile_pivot = (

    percentile_df

    .pivot_table(

        index=["company_id", "year"],

        columns="metric",

        values="percentile_rank"

    )

    .reset_index()

)

print("\nPercentile Pivot Shape:")

print(percentile_pivot.shape)

print(percentile_pivot.head())

# -----------------------------
# Merge Percentiles
# -----------------------------

report_df = report_df.merge(

    percentile_pivot,

    on=[
        "company_id",
        "year"
    ],

    how="left"

)

print("\nReport Shape After Merge:")

print(report_df.shape)


# -----------------------------
# Populate Worksheets
# -----------------------------

for group in peer_groups:

    ws = wb[group[:31]]

    group_df = report_df[
        report_df["peer_group_name"] == group
    ].copy()

    headers = report_columns + ["Benchmark"]

    # Header Row

    for col_num, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=1,
            column=col_num
        )

        cell.value = header
        cell.font = Font(bold=True)

    # Data Rows

    for row_num, (_, row) in enumerate(group_df.iterrows(), start=2):

        for col_num, column in enumerate(report_columns, start=1):

            if column in row.index:

                ws.cell(
                    row=row_num,
                    column=col_num
                ).value = row[column]

        ws.cell(
            row=row_num,
            column=len(report_columns) + 1
        ).value = row["is_benchmark"]

print("✅ Worksheets populated.")

# -----------------------------
# Cell Colors
# -----------------------------

green_fill = PatternFill(
    fill_type="solid",
    start_color="90EE90"
)

yellow_fill = PatternFill(
    fill_type="solid",
    start_color="FFF59D"
)

red_fill = PatternFill(
    fill_type="solid",
    start_color="FF9999"
)

gold_fill = PatternFill(
    fill_type="solid",
    start_color="FFD966"
)

# -----------------------------
# Percentile Columns
# -----------------------------

percentile_columns = [

    "ROE",
    "ROCE",
    "Net Profit Margin",
    "Debt to Equity",
    "Interest Coverage",
    "Asset Turnover",
    "Revenue CAGR 5Y",
    "PAT CAGR 5Y",
    "EPS CAGR 5Y"

]

# -----------------------------
# Apply Formatting
# -----------------------------

for ws in wb.worksheets:

    header = [
        cell.value
        for cell in ws[1]
    ]

    benchmark_col = header.index("Benchmark") + 1

    percentile_indexes = []

    for col in percentile_columns:

        if col in header:

            percentile_indexes.append(
                header.index(col) + 1
            )

    for row in range(2, ws.max_row + 1):

        # Highlight Benchmark Row

        if ws.cell(row, benchmark_col).value:

            for col in range(1, ws.max_column + 1):

                ws.cell(
                    row,
                    col
                ).fill = gold_fill

        # Percentile Colors

        for col in percentile_indexes:

            value = ws.cell(
                row,
                col
            ).value

            if value is None:
                continue

            if value >= 75:

                ws.cell(
                    row,
                    col
                ).fill = green_fill

            elif value <= 25:

                ws.cell(
                    row,
                    col
                ).fill = red_fill

            else:

                ws.cell(
                    row,
                    col
                ).fill = yellow_fill

# -----------------------------
# Median Row
# -----------------------------

for ws in wb.worksheets:

    median_row = ws.max_row + 2

    ws.cell(
        median_row,
        1
    ).value = "Median"

    ws.cell(
        median_row,
        1
    ).font = Font(bold=True)

    headers = [
        cell.value
        for cell in ws[1]
    ]

    for col in range(2, ws.max_column):

        values = []

        for row in range(2, median_row - 1):

            value = ws.cell(
                row,
                col
            ).value

            if isinstance(value, (int, float)):

                values.append(value)

        if values:

            ws.cell(
                median_row,
                col
            ).value = round(
                pd.Series(values).median(),
                2
            )

print("✅ Median rows added.")

# -----------------------------
# Save Workbook
# -----------------------------

import os

os.makedirs(
    "output",
    exist_ok=True
)

output_file = os.path.join(
    "output",
    "peer_comparison.xlsx"
)

wb.save(output_file)

print("\n✅ Peer Comparison Report Generated Successfully")

print(output_file)