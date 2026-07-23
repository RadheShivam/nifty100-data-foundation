import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from engine import ScreenerEngine
from presets import (
    QUALITY_COMPOUNDER,
    VALUE_PICK,
    GROWTH_ACCELERATOR,
    DIVIDEND_CHAMPION,
    DEBT_FREE_BLUECHIP,
    TURNAROUND_WATCH,
)


def main():

    presets = {
        "Quality Compounder": QUALITY_COMPOUNDER,
        "Value Pick": VALUE_PICK,
        "Growth Accelerator": GROWTH_ACCELERATOR,
        "Dividend Champion": DIVIDEND_CHAMPION,
        "Debt-Free Blue Chip": DEBT_FREE_BLUECHIP,
        "Turnaround Watch": TURNAROUND_WATCH,
    }

    os.makedirs("output", exist_ok=True)

    output_file = "output/screener_output.xlsx"

    # -----------------------------
    # Export all presets
    # -----------------------------
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        for sheet_name, config in presets.items():

            print(f"\nRunning: {sheet_name}")
            print(config)

            engine = ScreenerEngine(config=config)

            filtered = engine.apply_filters()

            filtered = filtered.sort_values(
                by=["sector_relative_score", "composite_quality_score"],
                ascending=[False, False],
            )

            filtered.to_excel(writer, sheet_name=sheet_name[:31], index=False)

            print(
                filtered[
                    [
                        "company_id",
                        "broad_sector",
                        "year",
                        "sector_relative_score",
                        "composite_quality_score",
                        "revenue_cagr_3yr",
                        "free_cash_flow_cr",
                    ]
                ]
            )

            print(f"Rows: {len(filtered)}")

    print(f"\n✅ Excel exported successfully: {output_file}")

    # -----------------------------
    # Colour Coding
    # -----------------------------

    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb = load_workbook(output_file)

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        headers = {}

        for cell in ws[1]:
            headers[cell.value] = cell.column

        for row in range(2, ws.max_row + 1):

            # ROE
            if "return_on_equity_pct" in headers:
                cell = ws.cell(row=row, column=headers["return_on_equity_pct"])
                if cell.value is not None:
                    cell.fill = green_fill if cell.value >= 15 else red_fill

            # Debt to Equity
            if "debt_to_equity" in headers:
                cell = ws.cell(row=row, column=headers["debt_to_equity"])
                if cell.value is not None:
                    cell.fill = green_fill if cell.value <= 2 else red_fill

            # Free Cash Flow
            if "free_cash_flow_cr" in headers:
                cell = ws.cell(row=row, column=headers["free_cash_flow_cr"])
                if cell.value is not None:
                    cell.fill = green_fill if cell.value > 0 else red_fill

            # Revenue CAGR 5yr
            if "revenue_cagr_5yr" in headers:
                cell = ws.cell(row=row, column=headers["revenue_cagr_5yr"])
                if cell.value is not None:
                    cell.fill = green_fill if cell.value >= 10 else red_fill

            # PAT CAGR 5yr
            if "pat_cagr_5yr" in headers:
                cell = ws.cell(row=row, column=headers["pat_cagr_5yr"])
                if cell.value is not None:
                    cell.fill = green_fill if cell.value >= 20 else red_fill

            # Dividend Payout
            if "dividend_payout_ratio_pct" in headers:
                cell = ws.cell(row=row, column=headers["dividend_payout_ratio_pct"])
                if cell.value is not None:
                    cell.fill = green_fill if cell.value <= 80 else red_fill

            # Revenue CAGR 3yr
            if "revenue_cagr_3yr" in headers:
                cell = ws.cell(row=row, column=headers["revenue_cagr_3yr"])
                if cell.value is not None:
                    cell.fill = green_fill if cell.value >= 10 else red_fill

    wb.save(output_file)

    print("✅ Colour coding applied successfully!")


if __name__ == "__main__":
    main()
